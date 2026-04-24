# """
# FastAPI wrapper around Advanced_RAG4.py
# Run:  uvicorn api:app --reload --port 8000
# """
#
# import sys
# import os
# from pathlib import Path
# from typing import Optional
# from contextlib import asynccontextmanager
#
# from fastapi import FastAPI, HTTPException
# from fastapi.middleware.cors import CORSMiddleware
# from fastapi.staticfiles import StaticFiles
# from fastapi.responses import FileResponse
# from pydantic import BaseModel
#
# # ── Make sure Advanced_RAG4 is importable ─────────────────────────────────────
# BASE_DIR = Path(__file__).parent
# sys.path.insert(0, str(BASE_DIR))
#
# from Advanced_RAG4 import (
#     load_documents_from_pdf,
#     load_documents_from_text,
#     create_embeddings_model,
#     split_documents_semantic,
#     split_documents_recursive,
#     build_faiss_vectorstore,
#     load_faiss_vectorstore,
#     create_hybrid_retriever,
#     load_reranker,
#     create_llm,
#     create_rag_prompt,
#     build_rag_chain,
#     ResponseCache,
#     ConversationMemory,
#     query_rag_pipeline,
#     setup_logger,
# )
#
# log = setup_logger("api.log")
#
# # ---------------------------------------------------------------------------
# # CONFIG  — edit these to match your setup
# # ---------------------------------------------------------------------------
#
# DOCUMENT_PATHS = [
#     r"C:\Users\SAISH PARKAR\PycharmProjects\RagStudyMaterial\articles\article.txt",
#     r"C:\Users\SAISH PARKAR\PycharmProjects\RagStudyMaterial\articles\company_employees_details.xlsx",
# ]
#
# EMBEDDINGS_MODEL = "all-MiniLM-L6-v2"
# LLM_MODEL        = "llama-3.3-70b-versatile"
# RERANKER_MODEL   = "cross-encoder/ms-marco-MiniLM-L-6-v2"
# FAISS_PATH       = str(BASE_DIR / "faiss_index")
# REBUILD_STORE    = True   # Set True on first run or when documents change
#
# # ---------------------------------------------------------------------------
# # GLOBAL PIPELINE STATE  (loaded once at startup)
# # ---------------------------------------------------------------------------
#
# pipeline: dict = {}
#
#
# def init_pipeline():
#     """Load all models and build / load the vector store once at startup."""
#     log.info("[API] Initialising RAG pipeline …")
#
#     embeddings = create_embeddings_model(EMBEDDINGS_MODEL)
#
#     # ── Load docs & build index (or just load existing index) ─────────────
#     if REBUILD_STORE:
#         all_docs = []
#         for path in DOCUMENT_PATHS:
#             p = Path(path)
#             if not p.exists():
#                 log.warning(f"[API] File not found, skipping: {path}")
#                 continue
#             if p.suffix.lower() == ".pdf":
#                 all_docs.extend(load_documents_from_pdf(str(p)))
#             elif p.suffix.lower() == ".txt":
#                 all_docs.extend(load_documents_from_text(str(p)))
#
#         if not all_docs:
#             raise RuntimeError("No documents loaded — check DOCUMENT_PATHS in api.py")
#
#         chunks = split_documents_semantic(all_docs, embeddings)
#         if len(chunks) < 5:
#             chunks = split_documents_recursive(all_docs)
#
#         vectorstore = build_faiss_vectorstore(chunks, embeddings, FAISS_PATH)
#     else:
#         vectorstore = load_faiss_vectorstore(embeddings, FAISS_PATH)
#         # Dummy chunk list — hybrid retriever needs real chunks; fall back to FAISS-only
#         chunks = None
#
#     # ── Retriever ─────────────────────────────────────────────────────────
#     if chunks:
#         retriever = create_hybrid_retriever(chunks, vectorstore, k=8)
#     else:
#         _faiss_ret = vectorstore.as_retriever(search_type="similarity", search_kwargs={"k": 8})
#         class _SimpleRetriever:
#             def __init__(self, r): self._r = r
#             def invoke(self, q):   return self._r.invoke(q)
#         retriever = _SimpleRetriever(_faiss_ret)
#         log.warning("[API] Loaded existing FAISS index — using FAISS-only retriever.")
#
#     reranker = load_reranker(RERANKER_MODEL)
#     llm      = create_llm(model_name=LLM_MODEL, temperature=0.0)
#     prompt   = create_rag_prompt()
#     chain    = build_rag_chain(llm, retriever, prompt)
#     cache    = ResponseCache(str(BASE_DIR / "rag_cache.json"))
#     memory   = ConversationMemory(max_turns=6)
#
#     pipeline.update(
#         chain=chain, retriever=retriever, reranker=reranker,
#         llm=llm, cache=cache, memory=memory,
#     )
#     log.info("[API] Pipeline ready ✅")
#
#
# # ---------------------------------------------------------------------------
# # LIFESPAN  (replaces deprecated @app.on_event)
# # ---------------------------------------------------------------------------
#
# @asynccontextmanager
# async def lifespan(app: FastAPI):
#     init_pipeline()
#     yield
#
#
# # ---------------------------------------------------------------------------
# # APP
# # ---------------------------------------------------------------------------
#
# app = FastAPI(
#     title="RAG Chatbot API",
#     description="Ask questions answered from your document corpus.",
#     version="1.0.0",
#     lifespan=lifespan,
# )
#
# app.add_middleware(
#     CORSMiddleware,
#     allow_origins=["*"],   # tighten in production
#     allow_methods=["*"],
#     allow_headers=["*"],
# )
#
# # Serve the chat UI from the same folder
# UI_DIR = BASE_DIR / "chat_ui"
# if UI_DIR.exists():
#     app.mount("/ui", StaticFiles(directory=str(UI_DIR), html=True), name="ui")
#
#
# # ---------------------------------------------------------------------------
# # SCHEMAS
# # ---------------------------------------------------------------------------
#
# class ChatRequest(BaseModel):
#     question: str
#     evaluate: bool = False          # enable RAGAS scoring (slower)
#     rewrite_query: bool = True
#
#
# class ChatResponse(BaseModel):
#     question: str
#     rewritten_query: Optional[str] = None
#     answer: str
#     sources: list[str]
#     eval_scores: Optional[dict] = None
#     latency_s: float
#     cached: bool
#
#
# class ResetResponse(BaseModel):
#     message: str
#
#
# # ---------------------------------------------------------------------------
# # ROUTES
# # ---------------------------------------------------------------------------
#
# @app.get("/", include_in_schema=False)
# def root():
#     index = BASE_DIR / "chat_ui" / "index.html"
#     if index.exists():
#         return FileResponse(str(index))
#     return {"message": "RAG API is running. POST /chat to ask a question."}
#
#
# @app.post("/chat", response_model=ChatResponse)
# def chat(req: ChatRequest):
#     if not req.question.strip():
#         raise HTTPException(status_code=400, detail="Question cannot be empty.")
#
#     try:
#         result = query_rag_pipeline(
#             chain           = pipeline["chain"],
#             retriever       = pipeline["retriever"],
#             question        = req.question,
#             reranker        = pipeline["reranker"],
#             llm             = pipeline["llm"],
#             memory          = pipeline["memory"],
#             cache           = pipeline["cache"],
#             evaluate        = req.evaluate,
#             rewrite_queries = req.rewrite_query,
#         )
#     except Exception as e:
#         log.error(f"[API] Pipeline error: {e}")
#         raise HTTPException(status_code=500, detail=str(e))
#
#     # Build deduplicated source list
#     sources = []
#     seen = set()
#     for doc in result.get("source_documents", []):
#         src  = doc.metadata.get("source", "unknown")
#         page = doc.metadata.get("page", "")
#         key  = f"{src}:{page}"
#         if key not in seen:
#             seen.add(key)
#             name = Path(src).name
#             sources.append(f"{name}" + (f" p.{page}" if page != "" else ""))
#
#     return ChatResponse(
#         question        = result["query"],
#         rewritten_query = result.get("rewritten_query"),
#         answer          = result["result"],
#         sources         = sources,
#         eval_scores     = result.get("eval_scores") or None,
#         latency_s       = round(result.get("latency_s", 0), 2),
#         cached          = result.get("cached", False),
#     )
#
#
# @app.post("/reset", response_model=ResetResponse)
# def reset():
#     """Clear conversation memory and response cache."""
#     pipeline["memory"].clear()
#     pipeline["cache"].clear()
#     return ResetResponse(message="Memory and cache cleared.")
#
#
# @app.get("/health")
# def health():
#     return {"status": "ok", "pipeline_loaded": bool(pipeline)}


"""
FastAPI wrapper around Advanced_RAG4.py
Run:  uvicorn api:app --reload --port 8000

Supports: .pdf  .txt  .xlsx
"""

import sys
import pickle
import time
from pathlib import Path
from typing import Optional
from contextlib import asynccontextmanager
from collections import defaultdict

from fastapi import FastAPI, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse
from pydantic import BaseModel

BASE_DIR = Path(__file__).parent
sys.path.insert(0, str(BASE_DIR))

from Advanced_RAG4 import (
    load_documents_from_pdf,
    load_documents_from_text,
    create_embeddings_model,
    split_documents_semantic,
    split_documents_recursive,
    build_faiss_vectorstore,
    load_faiss_vectorstore,
    create_hybrid_retriever,
    load_reranker,
    create_llm,
    create_rag_prompt,
    build_rag_chain,
    ResponseCache,
    ConversationMemory,
    rerank_documents,
    rewrite_query,
    format_docs,
    setup_logger,
)
from langchain_core.documents import Document
from langchain_core.output_parsers import StrOutputParser

log = setup_logger("api.log")

# ---------------------------------------------------------------------------
# CONFIG
# ---------------------------------------------------------------------------

DOCUMENT_PATHS = [
    r"C:\Users\SAISH PARKAR\PycharmProjects\RagStudyMaterial\articles\article.txt",
    r"C:\Users\SAISH PARKAR\PycharmProjects\RagStudyMaterial\articles\company_employees_details.xlsx",
]

EMBEDDINGS_MODEL      = "all-MiniLM-L6-v2"
LLM_MODEL             = "llama-3.3-70b-versatile"
RERANKER_MODEL        = "cross-encoder/ms-marco-MiniLM-L-6-v2"
FAISS_PATH            = str(BASE_DIR / "faiss_index")
CHUNKS_CACHE_PATH     = BASE_DIR / "chunks_cache.pkl"

REBUILD_STORE         = True   # ← Set True once when docs change, then False
RETRIEVAL_K           = 12
RERANK_TOP_N          = 6
MAX_CHUNKS_PER_SOURCE = 4

# ---------------------------------------------------------------------------
# XLSX LOADER
# ---------------------------------------------------------------------------

def load_documents_from_xlsx(file_path: str) -> list[Document]:
    """
    Convert every sheet of an Excel workbook into LangChain Documents.

    Strategy — two passes per sheet:
      1. HEADER ROW  → one Document describing the columns
      2. DATA ROWS   → one Document per row, formatted as
                       "Column: value | Column: value | …"
                       so both keyword (BM25) and semantic search work well.

    Each Document carries metadata:
        source  — file path
        sheet   — sheet name
        row     — 1-based row number  (data docs only)
        type    — "header" | "row"

    Requires: openpyxl  (pip install openpyxl)
    """
    try:
        import openpyxl
    except ImportError:
        raise RuntimeError(
            "openpyxl is required for xlsx support.\n"
            "Install it with:  pip install openpyxl"
        )

    docs: list[Document] = []
    wb = openpyxl.load_workbook(file_path, data_only=True)
    log.info(f"[XLSX] Loading '{file_path}' — sheets: {wb.sheetnames}")

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            log.warning(f"[XLSX] Sheet '{sheet_name}' is empty, skipping.")
            continue

        # ── detect header row ──────────────────────────────────────────────
        # First non-empty row is treated as the header
        header_row_idx = None
        headers: list[str] = []
        for i, row in enumerate(rows):
            non_empty = [str(c).strip() for c in row if c is not None and str(c).strip()]
            if non_empty:
                header_row_idx = i
                headers = [str(c).strip() if c is not None else f"col_{j}"
                           for j, c in enumerate(row)]
                break

        if header_row_idx is None:
            log.warning(f"[XLSX] Sheet '{sheet_name}' has no usable rows.")
            continue

        # ── header document ────────────────────────────────────────────────
        header_text = (
            f"Sheet: {sheet_name}\n"
            f"Columns: {', '.join(h for h in headers if h)}\n"
            f"Total rows: {len(rows) - header_row_idx - 1}"
        )
        docs.append(Document(
            page_content=header_text,
            metadata={"source": file_path, "sheet": sheet_name,
                      "row": 0, "type": "header"},
        ))

        # ── one document per data row ──────────────────────────────────────
        skipped = 0
        for row_idx, row in enumerate(rows[header_row_idx + 1:], start=1):
            # Skip completely empty rows
            if all(c is None or str(c).strip() == "" for c in row):
                skipped += 1
                continue

            # "Name: John | Age: 30 | Department: Engineering"
            pairs = []
            for col_name, cell_val in zip(headers, row):
                val = str(cell_val).strip() if cell_val is not None else ""
                if val and val.lower() != "none":
                    pairs.append(f"{col_name}: {val}")

            if not pairs:
                skipped += 1
                continue

            row_text = (
                f"Sheet: {sheet_name} | Row {row_idx}\n"
                + " | ".join(pairs)
            )
            docs.append(Document(
                page_content=row_text,
                metadata={"source": file_path, "sheet": sheet_name,
                          "row": row_idx, "type": "row"},
            ))

        log.info(
            f"[XLSX] Sheet '{sheet_name}': "
            f"{len(rows) - header_row_idx - 1 - skipped} data docs "
            f"({skipped} empty rows skipped)"
        )

    log.info(f"[XLSX] Total documents from '{Path(file_path).name}': {len(docs)}")
    return docs


# ---------------------------------------------------------------------------
# PER-SOURCE CAP
# ---------------------------------------------------------------------------

def cap_by_source(docs, max_per_source: int = MAX_CHUNKS_PER_SOURCE):
    counts: defaultdict = defaultdict(int)
    result = []
    for doc in docs:
        src = Path(doc.metadata.get("source", "unknown")).name
        if counts[src] < max_per_source:
            counts[src] += 1
            result.append(doc)
    return result


# ---------------------------------------------------------------------------
# PIPELINE INIT
# ---------------------------------------------------------------------------

def init_pipeline():
    log.info("[API] Initialising RAG pipeline …")
    embeddings = create_embeddings_model(EMBEDDINGS_MODEL)

    if REBUILD_STORE:
        all_docs: list[Document] = []

        for path in DOCUMENT_PATHS:
            p = Path(path)
            if not p.exists():
                log.warning(f"[API] File not found, skipping: {path}")
                continue

            suffix = p.suffix.lower()
            if suffix == ".pdf":
                all_docs.extend(load_documents_from_pdf(str(p)))
            elif suffix == ".txt":
                all_docs.extend(load_documents_from_text(str(p)))
            elif suffix in (".xlsx", ".xls"):
                all_docs.extend(load_documents_from_xlsx(str(p)))
            else:
                log.warning(f"[API] Unsupported file type, skipping: {path}")

        if not all_docs:
            raise RuntimeError("No documents loaded — check DOCUMENT_PATHS in api.py")

        log.info(f"[API] Total raw docs loaded: {len(all_docs)}")

        # Log per-source breakdown
        counts: defaultdict = defaultdict(int)
        for d in all_docs:
            counts[Path(d.metadata.get("source","?")).name] += 1
        for src, cnt in counts.items():
            log.info(f"[API]   {src}: {cnt} raw docs")

        # For xlsx row-docs (already small), skip semantic chunker — use them as-is
        # For pdf/txt, chunk normally
        xlsx_docs  = [d for d in all_docs if Path(d.metadata.get("source","")).suffix.lower() in (".xlsx",".xls")]
        other_docs = [d for d in all_docs if d not in xlsx_docs]

        chunks: list[Document] = list(xlsx_docs)   # xlsx rows stay whole

        if other_docs:
            text_chunks = split_documents_semantic(other_docs, embeddings)
            if len(text_chunks) < 5:
                log.warning("[API] Few semantic chunks — falling back to recursive splitter.")
                text_chunks = split_documents_recursive(other_docs)
            chunks.extend(text_chunks)

        log.info(f"[API] Total chunks for indexing: {len(chunks)}")
        counts2: defaultdict = defaultdict(int)
        for c in chunks:
            counts2[Path(c.metadata.get("source","?")).name] += 1
        for src, cnt in counts2.items():
            log.info(f"[API]   {src}: {cnt} chunks")

        vectorstore = build_faiss_vectorstore(chunks, embeddings, FAISS_PATH)

        with open(CHUNKS_CACHE_PATH, "wb") as f:
            pickle.dump(chunks, f)
        log.info(f"[API] Chunks persisted → {CHUNKS_CACHE_PATH}")

    else:
        vectorstore = load_faiss_vectorstore(embeddings, FAISS_PATH)
        chunks = None
        if CHUNKS_CACHE_PATH.exists():
            with open(CHUNKS_CACHE_PATH, "rb") as f:
                chunks = pickle.load(f)
            log.info(f"[API] Loaded {len(chunks)} chunks from cache.")
        else:
            log.warning("[API] No chunks_cache.pkl — set REBUILD_STORE=True once.")

    # ── Retriever ──────────────────────────────────────────────────────────
    if chunks:
        retriever = create_hybrid_retriever(chunks, vectorstore, k=RETRIEVAL_K)
    else:
        _f = vectorstore.as_retriever(search_type="similarity", search_kwargs={"k": RETRIEVAL_K})
        class _Wrap:
            def __init__(self, r): self._r = r
            def invoke(self, q):   return self._r.invoke(q)
        retriever = _Wrap(_f)
        log.warning("[API] FAISS-only retriever (no BM25).")

    reranker = load_reranker(RERANKER_MODEL)
    llm      = create_llm(model_name=LLM_MODEL, temperature=0.0)
    prompt   = create_rag_prompt()
    chain    = build_rag_chain(llm, retriever, prompt)
    cache    = ResponseCache(str(BASE_DIR / "rag_cache.json"))
    memory   = ConversationMemory(max_turns=6)

    pipeline.update(
        chain=chain, retriever=retriever, reranker=reranker,
        llm=llm, cache=cache, memory=memory, chunks=chunks,
    )
    log.info("[API] Pipeline ready ✅")


# ---------------------------------------------------------------------------
# LIFESPAN
# ---------------------------------------------------------------------------

pipeline: dict = {}

@asynccontextmanager
async def lifespan(app: FastAPI):
    init_pipeline()
    yield


# ---------------------------------------------------------------------------
# APP
# ---------------------------------------------------------------------------

app = FastAPI(
    title="RAG Chatbot API",
    description="Supports PDF, TXT, and XLSX documents.",
    version="3.0.0",
    lifespan=lifespan,
)

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

UI_DIR = BASE_DIR / "chat_ui"
if UI_DIR.exists():
    app.mount("/ui", StaticFiles(directory=str(UI_DIR), html=True), name="ui")


# ---------------------------------------------------------------------------
# SCHEMAS
# ---------------------------------------------------------------------------

class ChatRequest(BaseModel):
    question: str
    evaluate: bool = False
    rewrite_query: bool = True

class ChatResponse(BaseModel):
    question: str
    rewritten_query: Optional[str] = None
    answer: str
    sources: list[str]
    eval_scores: Optional[dict] = None
    latency_s: float
    cached: bool

class ResetResponse(BaseModel):
    message: str


# ---------------------------------------------------------------------------
# HELPERS
# ---------------------------------------------------------------------------

def build_source_list(docs) -> list[str]:
    seen, out = set(), []
    for doc in docs:
        src   = doc.metadata.get("source", "unknown")
        sheet = doc.metadata.get("sheet", "")
        row   = doc.metadata.get("row", "")
        # For xlsx show  filename › Sheet (row N)
        if sheet:
            key  = f"{src}:{sheet}:{row}"
            label = Path(src).name + f" › {sheet}" + (f" row {row}" if row else "")
        else:
            page = doc.metadata.get("page", "")
            key  = f"{src}:{page}"
            label = Path(src).name + (f" p.{page}" if page != "" else "")
        if key not in seen:
            seen.add(key)
            out.append(label)
    return out


# ---------------------------------------------------------------------------
# ROUTES
# ---------------------------------------------------------------------------

@app.get("/", include_in_schema=False)
def root():
    idx = BASE_DIR / "chat_ui" / "index.html"
    return FileResponse(str(idx)) if idx.exists() else {"message": "RAG API running."}


@app.post("/chat", response_model=ChatResponse)
def chat(req: ChatRequest):
    if not req.question.strip():
        raise HTTPException(status_code=400, detail="Question cannot be empty.")
    try:
        llm       = pipeline["llm"]
        retriever = pipeline["retriever"]
        reranker  = pipeline["reranker"]
        memory    = pipeline["memory"]
        cache     = pipeline["cache"]

        q        = rewrite_query(llm, req.question) if req.rewrite_query else req.question
        raw_docs = retriever.invoke(q)
        capped   = cap_by_source(raw_docs, MAX_CHUNKS_PER_SOURCE)
        final    = rerank_documents(reranker, q, capped, top_n=RERANK_TOP_N)
        ctx      = format_docs(final)

        cached_ans = cache.get(req.question, ctx) if cache else None
        if cached_ans:
            return ChatResponse(
                question=req.question, rewritten_query=q,
                answer=cached_ans, sources=build_source_list(final),
                latency_s=0.0, cached=True,
            )

        t0 = time.time()
        history = memory.format() if memory else ""
        chain   = create_rag_prompt() | llm | StrOutputParser()
        answer  = chain.invoke({
            "context":  ctx,
            "question": req.question,
            "history":  f"Conversation so far:\n{history}\n" if history else "",
        })
        elapsed = round(time.time() - t0, 2)

        if cache:  cache.set(req.question, ctx, answer)
        if memory: memory.add(req.question, answer)

        eval_scores = None
        if req.evaluate:
            from Advanced_RAG4 import evaluate_response
            eval_scores = evaluate_response(llm, req.question, answer, ctx)

    except Exception as e:
        log.error(f"[API] Error: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))

    return ChatResponse(
        question=req.question, rewritten_query=q,
        answer=answer, sources=build_source_list(final),
        eval_scores=eval_scores, latency_s=elapsed, cached=False,
    )


@app.post("/reset", response_model=ResetResponse)
def reset():
    pipeline["memory"].clear()
    pipeline["cache"].clear()
    return ResetResponse(message="Memory and cache cleared.")


@app.get("/health")
def health():
    chunks = pipeline.get("chunks")
    counts: defaultdict = defaultdict(int)
    if chunks:
        for c in chunks:
            counts[Path(c.metadata.get("source","?")).name] += 1
    return {
        "status": "ok",
        "pipeline_loaded": bool(pipeline),
        "total_chunks": len(chunks) if chunks else 0,
        "chunks_per_source": dict(counts),
    }